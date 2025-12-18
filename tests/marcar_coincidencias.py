"""
Script para marcar visualmente las coincidencias entre el archivo procesado automáticamente
y el archivo manual de Karol, usando colores en Excel.
"""
import sys
import os
import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Dict, Optional
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# Configurar codificación UTF-8 para Windows
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# Agregar el directorio raíz al path
sys.path.insert(0, str(Path(__file__).parent.parent))

def limpiar_valor_numerico(valor):
    """Limpia y convierte valores numéricos"""
    if pd.isna(valor) or valor == '' or valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    if isinstance(valor, str):
        valor = valor.replace('$', '').replace(',', '').replace(' ', '').replace('-', '')
        try:
            return float(valor)
        except:
            return 0.0
    return 0.0

def normalizar_texto(texto):
    """Normaliza texto para comparación"""
    if pd.isna(texto) or texto is None:
        return ''
    texto = str(texto).strip().upper()
    texto = ' '.join(texto.split())
    return texto

def calcular_score_coincidencia(manual: Dict, automatico: Dict) -> tuple:
    """
    Calcula el score de coincidencia entre un registro manual y automático.
    
    Returns:
        (score_porcentaje, coincidencias, diferencias)
    """
    columnas_comparar = [
        'justificacion',
        'nuevo_estado',
        'ratificar_grabar_diferencia',
        'observaciones'
    ]
    
    coincidencias = 0
    total = 0
    diferencias = []
    
    for col in columnas_comparar:
        if col in manual and col in automatico:
            total += 1
            valor_manual = normalizar_texto(manual.get(col, ''))
            valor_automatico = normalizar_texto(automatico.get(col, ''))
            
            if valor_manual == valor_automatico:
                coincidencias += 1
            else:
                diferencias.append({
                    'campo': col,
                    'manual': manual.get(col, ''),
                    'automatico': automatico.get(col, '')
                })
    
    score = (coincidencias / total * 100) if total > 0 else 0
    return score, coincidencias, diferencias

def crear_clave_registro(row, col_cajero: str) -> str:
    """Crea una clave única para identificar un registro"""
    cajero = str(row.get(col_cajero, ''))
    tipo = normalizar_texto(row.get('tipo_registro', ''))
    fecha = str(row.get('fecha_arqueo', '')) if 'fecha_arqueo' in row else ''
    return f"{cajero}_{tipo}_{fecha}"

def main(fecha_especifica: Optional[str] = None):
    """Función principal"""
    # Obtener fecha
    if fecha_especifica:
        fecha_str = fecha_especifica
    else:
        fecha_hoy = datetime.now()
        fecha_str = fecha_hoy.strftime('%d_%m_%Y')
    
    # Rutas de archivos
    proyecto_root = Path(__file__).parent.parent
    insumos_excel = proyecto_root / 'insumos_excel'
    
    # Buscar archivo manual
    patron_manual = f'gestion_{fecha_str}_ksgarro_Karol.xlsx'
    archivos_manuales = list(insumos_excel.glob(f'gestion_{fecha_str}_ksgarro_KAROL.xlsx'))
    if not archivos_manuales:
        archivos_manuales = list(insumos_excel.glob(f'gestion_{fecha_str}_ksgarro_Karol.xlsx'))
    
    if not archivos_manuales:
        print(f"✗ No se encontró el archivo manual: {patron_manual}")
        print(f"   Buscando en: {insumos_excel}")
        return
    
    ruta_manual = archivos_manuales[0]
    
    # Buscar archivo automático
    patron_automatico = f'gestion_{fecha_str}_ksgarro_procesado.xlsx'
    archivos_automaticos = list(insumos_excel.glob(patron_automatico))
    
    if not archivos_automaticos:
        print(f"✗ No se encontró el archivo automático: {patron_automatico}")
        return
    
    ruta_automatico = archivos_automaticos[0]
    
    print("=" * 100)
    print("MARCADO VISUAL DE COINCIDENCIAS")
    print("=" * 100)
    print(f"\nArchivo manual: {ruta_manual.name}")
    print(f"Archivo automático: {ruta_automatico.name}\n")
    
    # Cargar archivos
    try:
        df_manual = pd.read_excel(ruta_manual)
        df_automatico = pd.read_excel(ruta_automatico)
        print(f"✓ Archivos cargados correctamente")
        print(f"  - Manual: {len(df_manual)} registros")
        print(f"  - Automático: {len(df_automatico)} registros\n")
    except Exception as e:
        print(f"✗ Error al cargar archivos: {e}")
        return
    
    # Identificar columna de código de cajero
    col_cajero_manual = None
    col_cajero_automatico = None
    
    for col in ['codigo_cajero', 'NIT', 'cajero', 'CAJERO', 'CODIGO_CAJERO']:
        if col in df_manual.columns:
            col_cajero_manual = col
            break
    
    for col in ['codigo_cajero', 'NIT', 'cajero', 'CAJERO', 'CODIGO_CAJERO']:
        if col in df_automatico.columns:
            col_cajero_automatico = col
            break
    
    if not col_cajero_manual or not col_cajero_automatico:
        print("✗ No se pudo identificar la columna de código de cajero")
        return
    
    # Crear diccionario de registros manuales por clave
    registros_manuales = {}
    for idx, row in df_manual.iterrows():
        clave = crear_clave_registro(row, col_cajero_manual)
        registros_manuales[clave] = row.to_dict()
    
    # Agregar columnas de comparación al DataFrame automático
    df_resultado = df_automatico.copy()
    df_resultado['_score_coincidencia'] = 0.0
    df_resultado['_estado_coincidencia'] = ''
    df_resultado['_coincidencias'] = 0
    df_resultado['_total_campos'] = 0
    df_resultado['_diferencias_detalle'] = ''
    
    # Comparar cada registro automático con el manual
    for idx, row in df_automatico.iterrows():
        clave = crear_clave_registro(row, col_cajero_automatico)
        
        if clave in registros_manuales:
            registro_manual = registros_manuales[clave]
            score, coincidencias, diferencias = calcular_score_coincidencia(
                registro_manual, row.to_dict()
            )
            
            df_resultado.loc[idx, '_score_coincidencia'] = score
            df_resultado.loc[idx, '_coincidencias'] = coincidencias
            df_resultado.loc[idx, '_total_campos'] = len(diferencias) + coincidencias
            
            # Determinar estado
            if score == 100:
                estado = 'PERFECTO'
            elif score >= 50:
                estado = 'PARCIAL'
            else:
                estado = 'NO_COINCIDE'
            
            df_resultado.loc[idx, '_estado_coincidencia'] = estado
            
            # Detalle de diferencias
            if diferencias:
                detalle = ' | '.join([f"{d['campo']}: Manual='{d['manual']}' vs Auto='{d['automatico']}'" 
                                     for d in diferencias])
                df_resultado.loc[idx, '_diferencias_detalle'] = detalle
        else:
            df_resultado.loc[idx, '_estado_coincidencia'] = 'NO_EN_MANUAL'
            df_resultado.loc[idx, '_diferencias_detalle'] = 'Registro no encontrado en archivo manual'
    
    # Guardar archivo Excel con formato
    nombre_salida = f'gestion_{fecha_str}_ksgarro_procesado_CON_COLORES.xlsx'
    ruta_salida = insumos_excel / nombre_salida
    
    # Guardar primero como Excel
    df_resultado.to_excel(ruta_salida, index=False, engine='openpyxl')
    
    # Aplicar formato de colores
    wb = openpyxl.load_workbook(ruta_salida)
    ws = wb.active
    
    # Definir colores
    verde = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    amarillo = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    rojo = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    gris = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    
    # Encontrar columna de estado
    col_estado = None
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value == '_estado_coincidencia':
            col_estado = col_idx
            break
    
    # Aplicar colores según el estado
    if col_estado:
        for row_idx in range(2, ws.max_row + 1):
            estado_cell = ws.cell(row=row_idx, column=col_estado)
            estado = estado_cell.value
            
            # Aplicar color a toda la fila
            if estado == 'PERFECTO':
                color = verde
            elif estado == 'PARCIAL':
                color = amarillo
            elif estado == 'NO_COINCIDE':
                color = rojo
            else:  # NO_EN_MANUAL
                color = gris
            
            # Aplicar color a las columnas principales (hasta la columna de estado)
            for col_idx in range(1, col_estado + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = color
    
    # Ajustar ancho de columnas
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 20
    
    # Guardar archivo con formato
    wb.save(ruta_salida)
    
    # Estadísticas
    total_registros = len(df_resultado)
    perfectos = len(df_resultado[df_resultado['_estado_coincidencia'] == 'PERFECTO'])
    parciales = len(df_resultado[df_resultado['_estado_coincidencia'] == 'PARCIAL'])
    no_coinciden = len(df_resultado[df_resultado['_estado_coincidencia'] == 'NO_COINCIDE'])
    no_en_manual = len(df_resultado[df_resultado['_estado_coincidencia'] == 'NO_EN_MANUAL'])
    
    print("=" * 100)
    print("ESTADÍSTICAS DE COINCIDENCIA")
    print("=" * 100)
    print(f"\nTotal de registros: {total_registros}")
    print(f"✓ Coincidencias perfectas (VERDE): {perfectos} ({perfectos/total_registros*100:.1f}%)")
    print(f"⚠ Coincidencias parciales (AMARILLO): {parciales} ({parciales/total_registros*100:.1f}%)")
    print(f"✗ No coinciden (ROJO): {no_coinciden} ({no_coinciden/total_registros*100:.1f}%)")
    print(f"○ No encontrados en manual (GRIS): {no_en_manual} ({no_en_manual/total_registros*100:.1f}%)")
    
    print(f"\n✓ Archivo con colores guardado en: {ruta_salida}")
    print("\nLEYENDA DE COLORES:")
    print("  🟢 VERDE: Coincidencia perfecta (100%)")
    print("  🟡 AMARILLO: Coincidencia parcial (50-99%)")
    print("  🔴 ROJO: No coincide (<50%)")
    print("  ⚪ GRIS: Registro no encontrado en archivo manual")

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Marcar coincidencias con colores en Excel')
    parser.add_argument(
        '--fecha',
        type=str,
        help='Fecha específica en formato DD_MM_YYYY (ej: 15_12_2025) para analizar'
    )
    args = parser.parse_args()
    main(fecha_especifica=args.fecha)

